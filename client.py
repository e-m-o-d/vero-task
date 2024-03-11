#!/usr/bin/env python

import sys
import logging
import argparse
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Config
APP_NAME = 'VERO Python Task Client'
APP_VERSION = '1'
SERVER_URL = 'http://localhost:8000' 
VEHICLES_ENDPOINT = SERVER_URL + '/vehicles'
VEHICLES_CSV = 'vehicles.csv'
VEHICLES_XLS = 'vehicles_{isodate}.xlsx'
KEY_RNR = 'rnr'
KEY_HU = 'hu'
KEY_GRUPPE = 'gruppe'
KEY_LABELIDS = 'labelIds'
KEY_LABELCOLOR = '_labelColor'
COLOR_GREEN = '007500'
COLOR_ORANGE = 'FFA500'
COLOR_RED = 'B30000'

def processVehicleData(keys, colored):
    logging.info('Processing vehicles data')

    # Loading vehicles data
    logging.info('Loading vehicles data from Server')
    response = None
    try:
        with open(VEHICLES_CSV, 'rb') as csvfile:
            response = requests.post(VEHICLES_ENDPOINT, files={'file': csvfile})
    except Exception as err:
        logging.error(f'Server vehicles call failed unexpectedly: {err=}')
        return False
    if response.status_code != 200:
        logging.error(f'Server vehicles call failed: {response.json()["message"]}')
        return False
    logging.info(f'Server vehicles request: {response.json()["message"]}')

    vehicles = response.json()['vehicles']
    if not vehicles:
        logging.error(f'No vehicles in server response.')
        return False

    # Sorting vehicles by its 'gruppe' key (by specification)
    sortedVehicles = sorted(vehicles, key=lambda x: x[KEY_GRUPPE])

    # Keys filtering and sanity checks
    if not keys:
        keys = []
    else:
        # Filter out eventual wrong argument keys which does not exist in vehicles data
        # NOTE: also remove 'rnr' key - it is always included (hardcoded) in Excel table
        allVehicleKeys = sortedVehicles[0].keys()
        keys = list(filter(lambda k: ((k in allVehicleKeys) and (k != 'rnr')), keys))
        # Remove multiple key occurences, keeping keys order
        keys = list(dict.fromkeys(keys))

    # Finally create the Excel file
    if not createExcelFile(sortedVehicles, keys, colored):
        loging.error("Failed to create Excel file")
        return False

    return True

def createExcelFile(vehicles, keys, colored):
    logging.info('Creating Excel file')

    wb = Workbook()
    ws = wb.active

    # Add table header
    # NOTE: ASSUMPTION: This is not explicitly required by the task specification, but makes the table more readle
    header = [KEY_RNR]
    for k in keys:
        header.append(k)
    ws.append(header)
    ws.row_dimensions[1].font = Font(bold=True)
    if colored:
        ws.row_dimensions[1].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    # Find labelIds column index
    labelIdsIndex = None
    if KEY_LABELIDS in keys:
        labelIdsIndex = keys.index(KEY_LABELIDS) + 2 # columns are 1 based, +1 for the always first rnr column

    # Add table rows
    rowCount = 2 #rows are 1 based and header row is already added
    for v in vehicles:
        # add row containing hardcoded rnr key and the additional columns (keys)
        row = [v[KEY_RNR]]
        for k in keys:
            row.append(v[k])
        ws.append(row)

        # coloring options
        rowColor = None
        if colored:
            rowColor = colorByAge(v[KEY_HU])
            if rowColor:
                ws.row_dimensions[rowCount].fill = PatternFill(start_color=rowColor, end_color=rowColor, fill_type='solid')
        #if labelIdsIndex:
        if (labelIdsIndex and (KEY_LABELCOLOR in v) and (v[KEY_LABELCOLOR])):
            labelIdsColor = v[KEY_LABELCOLOR].lstrip('#')
            labelCell = ws.cell(row=rowCount, column=labelIdsIndex)
            labelCell.font = Font(color=labelIdsColor)
            if rowColor:
                # NOTE: this separate fill is needed, because settings cell style prevails the row style, incl. font (being it set before or after)
                labelCell.fill = PatternFill(start_color=rowColor, end_color=rowColor, fill_type='solid')

        rowCount += 1

    # Finally save the Excel file
    excelFilename = VEHICLES_XLS.format(isodate=datetime.now().strftime('%Y-%m-%d'))
    try:
        wb.save(excelFilename)
    except Exception as err:
        logging.error(f'Excel file {excelFilename} saving failed: {err=}')
        return False
    
    logging.info(f'Excel file saved: {excelFilename}')    
    return True

def dateDiffMonths(d1, d2):
    return (d1.year - d2.year) * 12 + d1.month - d2.month

def colorByAge(dateStr):
    monthsOld = dateDiffMonths(datetime.now(), datetime.strptime(dateStr, "%Y-%m-%d"))
    color = None
    if monthsOld > 0:
        if monthsOld <= 3:
            color = COLOR_GREEN
        elif monthsOld <= 12:
            color = COLOR_ORANGE
        else:
            color = COLOR_RED
    return color

if __name__ == '__main__':

    parser = argparse.ArgumentParser(prog=APP_NAME,
        description='Client app producing vehicles Excel file (vehicles_<iso-date>.xsls) based on CSV input (vehicles.csv) and active vehicles data from Baubuddy server')
    parser.add_argument('-k', '--keys', nargs='*', help="List of vehicle keys to be added as Excel columns. Example: -k gruppe kurzname info")
    parser.add_argument('-c', '--colored', action='store_true', help="Flag to turn on Excel table coloring")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
    
    logging.info(f'{APP_NAME} ({APP_VERSION})')
    logging.debug(f'    keys: {args.keys}')
    logging.debug(f'    colored: {args.colored}')

    if not processVehicleData(args.keys, args.colored):
        logging.error('Vehicles processing failed.')
        sys.exit(1)

    logging.info('Vehicles processing completed successfully.')
